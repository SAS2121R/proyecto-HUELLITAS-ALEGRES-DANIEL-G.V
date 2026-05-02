from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db.models import Count, Sum, F, Q

from xhtml2pdf import pisa
import io

from agenda.models import Cita
from historial.models import HistorialClinico
from productos.models import Producto
from servicios.models import Servicio
from mascotas.models import Mascota
from usuarios.decorators import role_required


def _render_to_pdf(template_path, context):
    """Renderiza una plantilla Django como respuesta HTTP en formato PDF."""
    html = render_to_string(template_path, context)
    buf = io.BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buf)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    buf.seek(0)
    return HttpResponse(buf.read(), content_type='application/pdf')


def _export_inventario_excel(queryset):
    """Exporta un queryset de Producto como respuesta HTTP en formato Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    headers = ['Nombre', 'Categoria', 'Precio', 'Stock', 'Stock Minimo', 'Proveedor', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row, p in enumerate(queryset, 2):
        estado = 'OK' if p.estado_stock == 'verde' else ('Bajo' if p.estado_stock == 'amarillo' else 'Critico')
        ws.cell(row=row, column=1, value=p.nombre)
        ws.cell(row=row, column=2, value=p.get_categoria_display())
        ws.cell(row=row, column=3, value=float(p.precio))
        ws.cell(row=row, column=4, value=p.cantidad_stock)
        ws.cell(row=row, column=5, value=p.stock_minimo)
        ws.cell(row=row, column=6, value=getattr(p.proveedor, 'nombre', '') or '-')
        ws.cell(row=row, column=7, value=estado)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
    return response


@login_required
@role_required('Administrador', 'Veterinario')
def reporte_citas(request):
    citas = Cita.objects.select_related('mascota', 'disponibilidad__veterinario').order_by('-disponibilidad__fecha')
    estado = request.GET.get('estado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    if estado:
        citas = citas.filter(estado=estado)
    if fecha_desde:
        citas = citas.filter(disponibilidad__fecha__gte=fecha_desde)
    if fecha_hasta:
        citas = citas.filter(disponibilidad__fecha__lte=fecha_hasta)
    total = citas.count()
    confirmadas = citas.filter(estado='confirmada').count()
    pendientes = citas.filter(estado='pendiente').count()
    canceladas = citas.filter(estado='cancelada').count()
    context = {
        'citas': citas,
        'total': total,
        'confirmadas': confirmadas,
        'pendientes': pendientes,
        'canceladas': canceladas,
    }
    return _render_to_pdf('reportes/reporte_citas.html', context)


@login_required
@role_required('Administrador', 'Veterinario')
def reporte_historial(request, mascota_pk):
    mascota = get_object_or_404(Mascota, pk=mascota_pk)
    historiales = HistorialClinico.objects.filter(mascota=mascota).select_related('veterinario').order_by('-fecha')
    context = {
        'mascota': mascota,
        'historiales': historiales,
    }
    return _render_to_pdf('reportes/reporte_historial.html', context)


@login_required
@role_required('Administrador', 'Veterinario')
def reporte_inventario(request):
    productos = Producto.objects.all().order_by('categoria', 'nombre')
    stock_bajo = request.GET.get('stock_bajo')
    if stock_bajo:
        productos = productos.filter(cantidad_stock__lt=5)
    total_productos = productos.count()
    total_valor = sum(p.precio * p.cantidad_stock for p in productos)
    critico = sum(1 for p in productos if p.estado_stock == 'rojo')
    bajo = sum(1 for p in productos if p.estado_stock == 'amarillo')
    context = {
        'productos': productos,
        'total_productos': total_productos,
        'total_valor': total_valor,
        'critico': critico,
        'bajo': bajo,
    }
    fmt = request.GET.get('format', 'pdf')
    if fmt == 'excel':
        return _export_inventario_excel(productos)
    return _render_to_pdf('reportes/reporte_inventario.html', context)


@login_required
@role_required('Administrador', 'Veterinario')
def reporte_servicios(request):
    servicios = Servicio.objects.all().order_by('categoria', 'nombre')
    categoria = request.GET.get('categoria')
    if categoria:
        servicios = servicios.filter(categoria=categoria)
    total = servicios.count()
    activos = servicios.filter(esta_activo=True).count()
    context = {
        'servicios': servicios,
        'total': total,
        'activos': activos,
    }
    return _render_to_pdf('reportes/reporte_servicios.html', context)


# ========================================
# ADMIN: Súper Reportes — Metrics Dashboard
# ========================================

@login_required
@role_required('Administrador')
def admin_metricas(request):
    """Panel de métricas del Administrador: Top 5 Productos, Productividad de Staff, Tasa de Cumplimiento."""
    context = _get_metricas_context(request)
    return render(request, 'reportes/admin_metricas.html', context)


# ========================================
# ADMIN: Exportar Métricas (PDF / Excel)
# ========================================

def _get_metricas_context(request):
    """Constructor de contexto compartido para las vistas de métricas."""
    from django.contrib.auth import get_user_model
    from entregas.models import Pedido, PedidoItem
    from usuarios.models import ConfiguracionClinica
    from django.utils import timezone

    today = timezone.now().date()
    start_of_month = today.replace(day=1)

    top_productos = PedidoItem.objects.filter(
        pedido__estado='entregado'
    ).values(
        'producto__nombre',
        'producto__categoria',
        'producto__precio',
    ).annotate(
        total_vendido=Sum('cantidad'),
        ingresos=Sum(F('producto__precio') * F('cantidad')),
        pedidos_count=Count('pedido', distinct=True),
    ).order_by('-total_vendido')[:5]

    Veterinario = get_user_model()
    vet_productividad = Veterinario.objects.filter(
        rol__nombre='Veterinario', is_active=True
    ).annotate(
        citas_mes=Count(
            'disponibilidades__citas',
            filter=Q(disponibilidades__fecha__gte=start_of_month),
            distinct=True,
        ),
        citas_total=Count('disponibilidades__citas', distinct=True),
    ).order_by('-citas_mes')

    total_entregados = Pedido.objects.filter(estado='entregado').count()
    con_incidente = Pedido.objects.filter(estado='cancelado').count()
    total_pedidos = Pedido.objects.count()
    tasa_cumplimiento = round((total_entregados / total_pedidos * 100) if total_pedidos > 0 else 0, 1)

    ingresos_mes = Pedido.objects.filter(
        estado='entregado', fecha_entrega__date__gte=start_of_month
    ).aggregate(total=Sum(F('items__producto__precio') * F('items__cantidad')))['total'] or 0

    config = ConfiguracionClinica.get_config()

    return {
        'top_productos': top_productos,
        'vet_productividad': vet_productividad,
        'total_entregados': total_entregados,
        'con_incidente': con_incidente,
        'total_pedidos': total_pedidos,
        'tasa_cumplimiento': tasa_cumplimiento,
        'ingresos_mes': ingresos_mes,
        'start_of_month': start_of_month,
        'today': today,
        'config': config,
    }


@login_required
def admin_metricas_pdf(request):
    """Exporta el panel de métricas como PDF."""
    context = _get_metricas_context(request)
    return _render_to_pdf('reportes/admin_metricas_pdf.html', context)


@login_required
def admin_metricas_excel(request):
    """Exporta el panel de métricas como archivo Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    context = _get_metricas_context(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Métricas de Negocio'

    # Estilo de encabezados
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    section_font = Font(bold=True, size=11, color='667eea')

    row = 1
    # Título
    ws.cell(row=row, column=1, value=f'Métricas de Negocio — {context["config"].nombre}')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ws.cell(row=row, column=1, value=f'Período: {context["start_of_month"].strftime("%Y-%m")} — {context["today"].strftime("%Y-%m-%d")}')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    # Tarjetas de resumen
    ws.cell(row=row, column=1, value='Resumen General')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    for label, val in [
        ('Pedidos Entregados', context['total_entregados']),
        ('Ingresos del Mes', f'${context["ingresos_mes"]}'),
        ('Incidentes', context['con_incidente']),
        ('Tasa de Cumplimiento', f'{context["tasa_cumplimiento"]}%'),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1
    row += 1

    # Top 5 Productos más vendidos
    ws.cell(row=row, column=1, value='Top 5 Productos Más Vendidos')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    for col, h in enumerate(['#', 'Producto', 'Categoría', 'Unidades', 'Ingresos'], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    for i, prod in enumerate(context['top_productos'], 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=prod['producto__nombre'])
        ws.cell(row=row, column=3, value=prod['producto__categoria'])
        ws.cell(row=row, column=4, value=prod['total_vendido'])
        ws.cell(row=row, column=5, value=float(prod['ingresos']))
        row += 1
    row += 1

    # Productividad de Staff
    ws.cell(row=row, column=1, value='Productividad de Staff')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    for col, h in ['Veterinario', 'Citas este Mes', 'Citas Total'], range(1, 4):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    # Corrección: usar zip correctamente para los encabezados
    for idx, h in enumerate(['Veterinario', 'Citas este Mes', 'Citas Total'], 1):
        cell = ws.cell(row=row, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    for vet in context['vet_productividad']:
        ws.cell(row=row, column=1, value=vet.get_full_name() or vet.email)
        ws.cell(row=row, column=2, value=vet.citas_mes)
        ws.cell(row=row, column=3, value=vet.citas_total)
        row += 1

    # Ajustar ancho de columnas automáticamente
    for col_idx in range(1, 6):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="metricas_negocio.xlsx"'
    return response
